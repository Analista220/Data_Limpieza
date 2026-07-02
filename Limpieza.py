import os
import json
import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import urllib.request
import urllib.error
import time
import io

# ─────────────────────────────────────────────────────────────────
#   CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────

CAMPOS_PLANTILLA = ["fecha", "vendedor", "cc", "referencia", "Valor Total", "cantidad", "cliente", "nit"]
RUTA_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config_distribuidores.json")
NULOS = ['nan', 'NAN', 'None', 'none', 'NONE', 'NaT', '']

# ─────────────────────────────────────────────────────────────────
#   API KEY
# ─────────────────────────────────────────────────────────────────


#xtfcxcgx

def pedir_api_key():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        root = tk.Tk()
        root.withdraw()
        mensaje = (
            "Creado por 🐐 Nicolas Tovar y 🐐 Cristian Morales\n\n"
            "Ingresa tu Anthropic API Key:\n"
            "(Solo se pide una vez por sesión)"
        )
        api_key = simpledialog.askstring("API Key requerida", mensaje, show="*")
    return api_key.strip() if api_key else None

# ─────────────────────────────────────────────────────────────────
#   DETECCIÓN DE FILA DE ENCABEZADOS (FILAS 1, 2 O 3)
# ─────────────────────────────────────────────────────────────────

def detectar_fila_encabezados(path: str) -> tuple[pd.DataFrame, int]:
    """
    Prueba las primeras 3 filas como posible encabezado.
    Una fila es válida si tiene al menos 3 columnas con nombre real (no Unnamed).
    Retorna (DataFrame completo con encabezados correctos, número de fila 1-based).
    """
    for fila_idx in range(3):
        try:
            df = pd.read_excel(path, header=fila_idx, dtype=str, nrows=5)
            df.columns = df.columns.astype(str).str.strip()
            cols_validas = [
                c for c in df.columns
                if c and not c.startswith("Unnamed:") and c.lower() != "nan"
            ]
            if len(cols_validas) >= 3:
                df_completo = pd.read_excel(path, header=fila_idx, dtype=str)
                df_completo.columns = df_completo.columns.astype(str).str.strip()
                df_completo = df_completo.loc[:, ~df_completo.columns.str.startswith("Unnamed:")]
                df_completo = df_completo.loc[:, df_completo.columns.str.lower() != "nan"]
                return df_completo, fila_idx + 1
        except Exception:
            continue

    # Fallback
    df_fallback = pd.read_excel(path, header=0, dtype=str)
    df_fallback.columns = df_fallback.columns.astype(str).str.strip()
    return df_fallback, 1

# ─────────────────────────────────────────────────────────────────
#   MAPEO CON IA
# ─────────────────────────────────────────────────────────────────

def mapear_con_ia(encabezados: list, nombre_archivo: str, api_key: str) -> tuple[dict, str] | tuple[None, None]:
    prompt = f"""Eres un asistente experto en normalización de datos de ventas.

Tengo un archivo Excel de un distribuidor llamado "{nombre_archivo}" con estos encabezados exactos:
{json.dumps(encabezados, ensure_ascii=False)}

Necesito dos cosas:

1. Mapear los encabezados a estos campos estándar:
- fecha        → columna con fechas de venta/documento
- vendedor     → columna con nombre o código del vendedor/asesor
- cc           → columna con cédula o identificación del vendedor (puede no existir)
- referencia   → columna con el nombre DESCRIPTIVO del producto (texto, no código numérico). Si hay nombre y código de producto, elige el nombre.
- Valor Total  → columna con el valor/precio total de la venta (no el costo, no descuentos)
- cantidad     → columna con cantidad de unidades o cajas vendidas
- cliente      → columna con nombre descriptivo del cliente (texto, no código). Si hay nombre y código, elige el nombre.
- nit          → columna con NIT o identificación numérica del cliente (puede no existir)

2. Identificar un nombre corto y estable del distribuidor a partir del nombre del archivo, ignorando palabras como meses (enero, febrero, marzo, abril, mayo, junio, julio, agosto, septiembre, octubre, noviembre, diciembre), años (2024, 2025, 2026), palabras genéricas (ventas, liquidacion, formato, planilla, reporte, cierre, corregido) y números de versión. El nombre corto debe ser la palabra o palabras que identifican ÚNICAMENTE a ese distribuidor.

Ejemplos de nombre corto:
- "Ventas GMJ abril 2026" → "GMJ"
- "LIQUIDACION TITANES MABE CIERRE ABRIL 2026-KIRAMAR" → "KIRAMAR"
- "VENTA MULTIELECTO ABRIL 2026" → "MULTIELECTO"
- "GMJ CORREGIDO Formato Planilla de ventas 2026 MABE abril2026" → "GMJ"
- "agaval Ventas_Mabe_2026-abril" → "agaval"

Reglas del mapeo:
1. Cada valor del mapeo debe ser UNA LISTA con el encabezado exacto del Excel.
2. Si un campo no tiene columna clara, usa lista vacía: [].
3. Si la columna de vendedor parece agrupada (pocos valores únicos repetidos), agrega "rellenar_vendedor": true.
4. Para "referencia" y "cliente": si hay nombre + código, SIEMPRE prioriza el nombre en texto.

Responde ÚNICAMENTE con un objeto JSON válido con esta estructura exacta, sin explicaciones ni bloques de código:
{{"_clave": "NOMBRE_CORTO_AQUI", "fecha": [...], "vendedor": [...], "cc": [...], "referencia": [...], "Valor Total": [...], "cantidad": [...], "cliente": [...], "nit": [...], "rellenar_vendedor": false}}"""

    body = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 500,
        "messages": [{"role": "user", "content": prompt}]
    }).encode("utf-8")

    def construir_req():
        return urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=body,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01"
            },
            method="POST"
        )

    req = construir_req()

    for intento in range(3):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                respuesta = json.loads(resp.read().decode("utf-8"))
                texto = respuesta["content"][0]["text"].strip()
                texto = texto.replace("```json", "").replace("```", "").strip()
                mapeo = json.loads(texto)

                clave_corta = mapeo.pop("_clave", None)
                if not clave_corta:
                    clave_corta = "_".join(nombre_archivo.split()[:2])
                clave_corta = clave_corta.strip().upper()

                print(f"   🤖 IA generó mapeo | Clave del distribuidor: '{clave_corta}'")
                return mapeo, clave_corta

        except urllib.error.HTTPError as e:
            cuerpo = e.read().decode("utf-8")
            if e.code == 429:
                espera = 20 * (intento + 1)
                print(f"   ⏳ Límite de velocidad. Esperando {espera}s...")
                time.sleep(espera)
                req = construir_req()
                continue
            print(f"   ❌ Error HTTP {e.code}: {cuerpo}")
            return None, None
        except Exception as e:
            print(f"   ❌ Error al llamar a la IA: {e}")
            return None, None

    print("   ❌ Se agotaron los 3 reintentos.")
    return None, None

# ─────────────────────────────────────────────────────────────────
#   JSON
# ─────────────────────────────────────────────────────────────────

def cargar_json() -> dict:
    if os.path.exists(RUTA_JSON):
        with open(RUTA_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_json(config: dict):
    with open(RUTA_JSON, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    print(f"   💾 Mapeo guardado en config_distribuidores.json")

# ─────────────────────────────────────────────────────────────────
#   RESOLUCIÓN DE MAPEO
# ─────────────────────────────────────────────────────────────────

def resolver_mapeo(nombre_archivo: str, nombre_carpeta: str, encabezados: list, config: dict, api_key: str):
    texto_busqueda = f"{nombre_archivo.upper()} {nombre_carpeta.upper()}"

    # 1. Buscar en JSON existente
    for llave, datos in config.items():
        if llave.upper() in ("GENERAL",):
            continue
        if llave.upper() in texto_busqueda:
            print(f"   🎯 Mapeo encontrado en JSON para: '{llave}'")
            return datos, datos.get("rellenar_vendedor", False)

    # 2. Bloque GENERAL como fallback
    if "GENERAL" in config:
        general = config["GENERAL"]
        coincidencias = sum(
            1 for sinonimos in general.values()
            if isinstance(sinonimos, list) and any(s in encabezados for s in sinonimos)
        )
        if coincidencias >= 2:
            print(f"   ℹ️  Usando bloque GENERAL (coincidencias: {coincidencias})")
            return general, general.get("rellenar_vendedor", False)

    # 3. Llamada a la IA
    if api_key:
        print(f"   🤖 No se encontró mapeo local. Consultando a Claude...")
        nuevo_mapeo, clave_corta = mapear_con_ia(encabezados, nombre_archivo, api_key)
        if nuevo_mapeo and clave_corta:
            config[clave_corta] = nuevo_mapeo
            guardar_json(config)
            return nuevo_mapeo, nuevo_mapeo.get("rellenar_vendedor", False)

    print(f"   ⚠️  No se pudo obtener mapeo para {nombre_archivo}. Se saltará.")
    return None, False

# ─────────────────────────────────────────────────────────────────
#   PROCESAMIENTO PRINCIPAL
# ─────────────────────────────────────────────────────────────────

def cargar_datos_distribuidores():
    root = tk.Tk()
    root.withdraw()

    file_paths = filedialog.askopenfilenames(
        title="Selecciona los archivos de los distribuidores",
        filetypes=[("Archivos de Excel", "*.xlsx *.xls")]
    )
    if not file_paths:
        print("No se seleccionaron archivos.")
        return

    template_path = filedialog.askopenfilename(
        title="Selecciona el archivo plantilla",
        filetypes=[("Archivos de Excel", "*.xlsx *.xls")]
    )
    if not template_path:
        print("No se seleccionó plantilla.")
        return

    api_key = pedir_api_key()
    if not api_key:
        print("⚠️  Sin API key. Solo se usarán mapeos del JSON existente.")

    config = cargar_json()
    carpeta_salida = os.path.join(os.path.dirname(template_path), "Distribuciones_Procesadas")
    os.makedirs(carpeta_salida, exist_ok=True)

    exitosos, fallidos = [], []

    for path in file_paths:
        nombre_archivo = os.path.basename(path)
        nombre_carpeta = os.path.basename(os.path.dirname(path))

        print(f"\n{'─'*60}")
        print(f"📄 Procesando: {nombre_archivo}  (carpeta: {nombre_carpeta})")

        try:
            # ── 1. Detectar fila de encabezados ──────────────────
            df_origen, fila_detectada = detectar_fila_encabezados(path)
            print(f"   📋 Encabezados detectados en fila {fila_detectada}: {list(df_origen.columns)}")

            encabezados = list(df_origen.columns)

            # ── 2. Obtener mapeo ──────────────────────────────────
            mapeo_sinonimos, debe_rellenar = resolver_mapeo(
                nombre_archivo, nombre_carpeta, encabezados, config, api_key
            )

            if mapeo_sinonimos is None:
                fallidos.append(nombre_archivo)
                continue

            # ── 3. Construir mapeo con comparación exacta + case-insensitive ──
            columnas_a_extraer = {}
            campos_sin_columna = []

            for campo_plantilla, sinonimos in mapeo_sinonimos.items():
                if campo_plantilla == "rellenar_vendedor":
                    continue
                if not isinstance(sinonimos, list) or len(sinonimos) == 0:
                    campos_sin_columna.append(campo_plantilla)
                    continue

                encontrado = False
                for sinonimo in sinonimos:
                    # Comparación exacta primero
                    if sinonimo in df_origen.columns:
                        columnas_a_extraer[sinonimo] = campo_plantilla
                        encontrado = True
                        break
                    # Comparación case-insensitive como fallback
                    coincidencia = next(
                        (c for c in df_origen.columns if c.strip().lower() == sinonimo.strip().lower()),
                        None
                    )
                    if coincidencia:
                        columnas_a_extraer[coincidencia] = campo_plantilla
                        encontrado = True
                        break

                if not encontrado:
                    campos_sin_columna.append(campo_plantilla)

            if campos_sin_columna:
                print(f"   ⚠️  Campos sin columna encontrada: {campos_sin_columna}")

            if not columnas_a_extraer:
                print(f"   ❌ Ninguna columna coincidió. Saltando...")
                fallidos.append(nombre_archivo)
                continue

            print(f"   🔗 Columnas mapeadas: {columnas_a_extraer}")

            # ── 4. Extraer, renombrar y limpiar filas vacías ──────
            df_filtrado = df_origen[list(columnas_a_extraer.keys())].rename(columns=columnas_a_extraer)
            df_filtrado = df_filtrado.dropna(how="all").reset_index(drop=True)

            # ── 5. Cargar plantilla y alinear columnas ────────────
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            columnas_finales = [cell.value for cell in ws[1] if cell.value is not None]

            for col in columnas_finales:
                if col not in df_filtrado.columns:
                    df_filtrado[col] = None

            df_final = df_filtrado[columnas_finales].copy()

            # ── 6. Limpiezas y conversiones ───────────────────────
            if 'vendedor' in df_final.columns:
                df_final['vendedor'] = df_final['vendedor'].replace(NULOS, None)
                if debe_rellenar:
                    print("   ⬇️  Aplicando fill-down en columna 'vendedor'.")
                    df_final['vendedor'] = df_final['vendedor'].ffill()

            if 'cantidad' in df_final.columns:
                df_final['cantidad'] = pd.to_numeric(df_final['cantidad'], errors='coerce').fillna(0).round(0).astype(int)

            if 'Valor Total' in df_final.columns:
                df_final['Valor Total'] = pd.to_numeric(df_final['Valor Total'], errors='coerce').fillna(0).round(0).astype(int)

            if 'fecha' in df_final.columns:
                df_final['fecha'] = df_final['fecha'].replace(NULOS, None)

            if 'cc' in df_final.columns:
                df_final['cc'] = df_final['cc'].fillna('Sin cc').replace(NULOS, 'Sin cc')

            if 'nit' in df_final.columns:
                df_final['nit'] = df_final['nit'].fillna('Sin nit').replace(NULOS, 'Sin nit')

            # ── 7. Estilos desde la plantilla ─────────────────────
            estilos_columnas = {}
            if ws.max_row >= 2:
                for col_idx in range(1, ws.max_column + 1):
                    celda = ws.cell(row=2, column=col_idx)
                    estilos_columnas[col_idx] = {
                        'font': copy(celda.font), 'border': copy(celda.border),
                        'fill': copy(celda.fill), 'number_format': celda.number_format,
                        'alignment': copy(celda.alignment)
                    }
                ws.delete_rows(2, amount=ws.max_row)

            for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    nueva_celda = ws.cell(row=r_idx, column=c_idx, value=value)
                    if c_idx in estilos_columnas:
                        for prop in ('font', 'border', 'fill', 'number_format', 'alignment'):
                            if estilos_columnas[c_idx][prop]:
                                setattr(nueva_celda, prop, estilos_columnas[c_idx][prop])

            # ── 8. Limpiar última fila ────────────────────────────
            ultima_fila = ws.max_row
            if ultima_fila >= 2:
                indices_columnas = {nombre: idx + 1 for idx, nombre in enumerate(columnas_finales)}
                for campo in ['fecha', 'cc', 'nit', 'vendedor', 'cliente', 'referencia']:
                    if campo in indices_columnas:
                        ws.cell(row=ultima_fila, column=indices_columnas[campo], value=None)

            output_path = os.path.join(carpeta_salida, f"Limpio_{nombre_archivo}")
            wb.save(output_path)
            print(f"   ✔️  Guardado: Limpio_{nombre_archivo}")
            exitosos.append(nombre_archivo)

        except Exception as e:
            print(f"   ❌ Error en {nombre_archivo}: {e}")
            fallidos.append(nombre_archivo)

    print(f"\n{'═'*60}")
    print(f"✅ Procesados exitosamente: {len(exitosos)}")
    for f in exitosos:
        print(f"   • {f}")
    if fallidos:
        print(f"\n❌ Fallidos ({len(fallidos)}):")
        for f in fallidos:
            print(f"   • {f}")
    print(f"\n📁 Archivos guardados en:\n   {carpeta_salida}")


if __name__ == "__main__":
    print("╔══════════════════════════════════════════════════════════╗")
    print("║         SISTEMA DE LIQUIDACIÓN DE DISTRIBUIDORES         ║")
    print("║       Creado por Nicolas Tovar y Cristian Morales        ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print()
    cargar_datos_distribuidores()